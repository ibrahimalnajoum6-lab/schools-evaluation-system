import streamlit as st
import sqlite3
import pandas as pd
import os
import json
import io
import base64
import requests
import threading
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit.components.v1 as components

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

# -------------------------------------------------------------
# إعداد رابط Google Apps Script للرفع المباشر
# -------------------------------------------------------------
APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbzbXt7ZJ1qjdnES24kGMDTYifU9MG3eKQRbH3nRu-QUz1Nk2cHvJnSVatZEd2noYARs/exec"

def upload_file_to_drive(file_bytes, filename, school_name, visit_date_obj, mime_type='application/octet-stream'):
    try:
        month_folder_name = f"{visit_date_obj.month}-{visit_date_obj.year}"
        encoded_file = base64.b64encode(file_bytes).decode('utf-8')
        
        payload = {
            "fileName": filename,
            "fileBytes": encoded_file,
            "mimeType": mime_type,
            "schoolName": school_name,
            "monthYear": month_folder_name
        }
        
        response = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        res_data = response.json()
        
        if res_data.get("status") == "success":
            return res_data.get("url")
        else:
            return None
    except Exception:
        return None

def auto_backup_database_to_drive():
    try:
        db_path = "evaluation_system.db"
        if os.path.exists(db_path):
            with open(db_path, "rb") as f:
                db_bytes = f.read()
            backup_filename = f"DB_Backup_Schools_System_{date.today()}_{int(datetime.now().timestamp())}.db"
            upload_file_to_drive(
                db_bytes,
                backup_filename,
                "النسخ_الاحتياطية_للنظام",
                date.today(),
                "application/x-sqlite3"
            )
    except Exception:
        pass

def background_upload_task(eval_id, eval_data_dict, uploaded_files_data, school_name, visit_date):
    """مهمة معالجة خلفية لرفع الملفات والنسخ دون تجميد واجهة المستخدم"""
    try:
        drive_links = []
        excel_bytes = generate_evaluation_excel_form(eval_data_dict)
        excel_link = upload_file_to_drive(
            excel_bytes, f"استمارة_{eval_data_dict['teacher_name']}_{visit_date}.xlsx", school_name, visit_date,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        if excel_link:
            drive_links.append(excel_link)

        for f_name, f_bytes, f_type in uploaded_files_data:
            link = upload_file_to_drive(f_bytes, f_name, school_name, visit_date, f_type)
            if link:
                drive_links.append(link)

        if drive_links:
            conn = sqlite3.connect("evaluation_system.db")
            c = conn.cursor()
            c.execute("UPDATE evaluations SET drive_links=? WHERE id=?", (",".join(drive_links), eval_id))
            conn.commit()
            conn.close()

        auto_backup_database_to_drive()
    except Exception:
        pass

def delete_evaluation_by_id(eval_id):
    conn = sqlite3.connect("evaluation_system.db")
    c = conn.cursor()
    c.execute("DELETE FROM evaluations WHERE id=?", (eval_id,))
    conn.commit()
    conn.close()
    threading.Thread(target=auto_backup_database_to_drive).start()

# -------------------------------------------------------------
# إعداد الصفحة وتصاميم الـ CSS (مع إخفاء شارة Streamlit تماماً)
# -------------------------------------------------------------
st.set_page_config(
    page_title="منظومة تقييم المدارس الشرعية",
    page_icon="🕌",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800;900&display=swap');
    
    html, body, [class*="css"] {
        direction: rtl;
        text-align: right;
        font-family: 'Tajawal', sans-serif !important;
        scroll-behavior: smooth;
        background-color: #f8fafc;
        font-size: 16px;
    }
    
    [data-testid="stSidebar"], [data-testid="stSidebarCollapseButton"], [data-testid="collapsedControl"], header[data-testid="stHeader"] {
        display: none !important;
    }
    
    /* إخفاء شارة Streamlit والزر الأحمر السفلي نهائياً */
    #MainMenu {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    div[class*="viewerBadge"], 
    [data-testid="stStatusWidget"],
    .viewerBadge_container__1QSob {
        display: none !important;
        visibility: hidden !important;
    }
    
    .block-container {
        padding: 0.8rem !important;
        max-width: 100% !important;
    }
    
    .mobile-card {
        background-color: #ffffff;
        border-radius: 18px;
        padding: 20px;
        margin-bottom: 16px;
        border: 2px solid #e2e8f0;
        box-shadow: 0 4px 16px rgba(0,0,0,0.05);
    }
    
    .score-banner {
        background: linear-gradient(135deg, #0d5c3a 0%, #15803d 100%);
        color: white;
        border-radius: 18px;
        padding: 20px;
        text-align: center;
        box-shadow: 0 8px 22px rgba(13,92,58,0.28);
        margin-bottom: 18px;
    }

    .success-box {
        background: linear-gradient(135deg, #065f46 0%, #047857 100%);
        color: white;
        border-radius: 18px;
        padding: 22px;
        text-align: center;
        box-shadow: 0 8px 24px rgba(6, 95, 70, 0.35);
        margin-bottom: 20px;
        border: 2px solid #34d399;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: #e2e8f0;
        padding: 8px;
        border-radius: 16px;
        margin-bottom: 16px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 12px;
        padding: 12px 18px;
        font-weight: 800;
        font-size: 16px;
        color: #334155;
        background-color: transparent;
        border: none !important;
        transition: all 0.25s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #ffffff !important;
        color: #0d5c3a !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.12);
    }
    
    .criterion-item {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-right: 6px solid #0d5c3a;
        border-radius: 12px;
        padding: 14px 16px;
        margin-bottom: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.03);
    }
    
    .stButton > button {
        border-radius: 14px !important;
        font-weight: 800 !important;
        font-size: 17px !important;
        padding: 14px 22px !important;
        min-height: 52px !important;
        box-shadow: 0 3px 8px rgba(0,0,0,0.08) !important;
    }
    
    input, select, textarea {
        font-size: 16px !important;
        padding: 12px !important;
    }
    
    .stRadio [role="radiogroup"] {
        gap: 12px !important;
    }
</style>
""", unsafe_allow_html=True)

# -------------------------------------------------------------
# قاعدة البيانات
# -------------------------------------------------------------
def init_db():
    conn = sqlite3.connect("evaluation_system.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password TEXT, full_name TEXT, specialization TEXT, role TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS schools (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, gender TEXT, location TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS evaluations (
        id INTEGER PRIMARY KEY AUTOINCREMENT, committee_no TEXT, visit_date TEXT, academic_year TEXT, semester TEXT,
        school_name TEXT, gender_type TEXT, supervisor_name TEXT, teacher_name TEXT, subject TEXT,
        specialization TEXT, student_count INTEGER, grade_level TEXT, section TEXT, lesson_topic TEXT,
        job_status TEXT, experience TEXT, scores_json TEXT, total_score INTEGER, rating TEXT,
        excellence_points TEXT, dev_points TEXT, suggestions TEXT, media_paths TEXT, drive_links TEXT, status TEXT
    )''')
    
    c.execute("PRAGMA table_info(evaluations)")
    columns = [col[1] for col in c.fetchall()]
    if "drive_links" not in columns:
        c.execute("ALTER TABLE evaluations ADD COLUMN drive_links TEXT")

    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        supervisors = [
            ("admin", "admin123", "مدير التعليم الشرعي", "إدارة", "Admin"),
            ("ibrahim", "123456", "إبراهيم احمد النجوم", "لغة إنكليزية", "Supervisor"),
            ("m_shabo", "123456", "محمد مصطفى المصطفى الشعبو", "شريعة", "Supervisor"),
            ("shadi_h", "123456", "شادي أحمد حلاق", "شريعة", "Supervisor"),
            ("m_hout", "123456", "محمد حسن الحوت", "شريعة", "Supervisor"),
            ("abdullah_a", "123456", "عبد الله عارف", "شريعة", "Supervisor"),
            ("m_muhyiddin", "123456", "محمد محي الدين محمد", "شريعة", "Supervisor"),
            ("waddah_m", "123456", "وضاح محمد لمعت مخللاتي", "شريعة", "Supervisor"),
            ("amina_y", "123456", "أمينة يوسف العبد الله", "شريعة", "Supervisor"),
            ("muna_h", "123456", "منى جمعة الحسن", "شريعة", "Supervisor"),
            ("hasnaa_h", "123456", "حسناء حسن الحاج إبراهيم", "شريعة", "Supervisor"),
            ("muna_haj", "123456", "منى حسن الحاج إبراهيم", "شريعة", "Supervisor"),
            ("abdulqader_a", "123456", "عبد القادر محمد شحادة الأحمد", "لغة عربية", "Supervisor"),
            ("ruba_m", "123456", "ربى ماهر مكتبي", "لغة عربية", "Supervisor"),
            ("muthanna_h", "123456", "المثنى محمود الدياب الحماده", "رياضيات", "Supervisor"),
            ("baraa_s", "123456", "براءه محمد جهاد سيلم", "رياضيات", "Supervisor"),
            ("m_masto", "123456", "محمد حسين مسطو", "علم أحياء", "Supervisor"),
            ("rana_m", "123456", "رنا عبد الحميد معرستاوي", "علم أحياء", "Supervisor"),
            ("m_omran", "123456", "محمد أمين عبد الله عمران", "فيزياء وكيمياء", "Supervisor"),
            ("bushra_e", "123456", "بشرى محمود عيد", "فيزياء وكيمياء", "Supervisor"),
            ("hossam_r", "123456", "حسام عمر رسلان", "فلسفة", "Supervisor"),
            ("aya_z", "123456", "آية أنور زيتاني", "فلسفة", "Supervisor"),
            ("abdulrahim_d", "123456", "عبد الرحيم محمد دلو", "جغرافيا", "Supervisor"),
            ("shahla_s", "123456", "شهلا محمد شيخوني", "جغرافيا", "Supervisor"),
            ("m_rajab", "123456", "محمد مصطفى الرجب", "تاريخ", "Supervisor"),
            ("nafisa_f", "123456", "نفيسة فارس الفارس", "لغة إنكليزية", "Supervisor")
        ]
        c.executemany("INSERT INTO users (username, password, full_name, specialization, role) VALUES (?, ?, ?, ?, ?)", supervisors)

    c.execute("SELECT COUNT(*) FROM schools")
    if c.fetchone()[0] == 0:
        schools_data = [
            ("الثانوية الخسروية", "ذكور", "بجانب القلعة"),
            ("الثانوية الشرعية الأولى (العرقوب)", "ذكور", "العرقوب"),
            ("ثانوية صلاح الدين", "ذكور", "صلاح الدين"),
            ("ثانوية الصاخور الشرعية للبنين", "ذكور", "الصاخور"),
            ("ثانوية الصاخور الشرعية للبنات", "إناث", "الصاخور"),
            ("الثانوية النعمانية الشرعية للبنين", "ذكور", "جمعية الزهراء"),
            ("الثانوية النعمانية الشرعية للبنات", "إناث", "جمعية الزهراء"),
            ("ثانوية هارون الرشيد الشرعية للبنين", "ذكور", "الراموسة"),
            ("ثانوية هارون الرشيد الشرعية للبنات", "إناث", "الراموسة"),
            ("ثانوية معروف الكرخي للبنين", "ذكور", "الشيخ مقصود"),
            ("ثانوية معروف الكرخي للبنات", "إناث", "الشيخ مقصود"),
            ("ثانوية المنصورة الشرعية للبنين", "ذكور", "المنصورة"),
            ("ثانوية المنصورة الشرعية للبنات", "إناث", "المنصورة"),
            ("ثانوية الإرث النبوي للبنين", "ذكور", "باب النيرب"),
            ("ثانوية الإرث النبوي للبنات", "إناث", "المدرسة العلمية"),
            ("ثانوية الحمدانية الشرعية للبنين", "ذكور", "الحمدانية"),
            ("ثانوية ابن الجوزي الشرعية للبنين", "ذكور", "جمعية الريادة"),
            ("ثانوية ابن الجوزي الشرعية للبنات", "إناث", "جمعية الريادة"),
            ("ثانوية عز العلماء للبنين", "ذكور", "المشارقة"),
            ("ثانوية زين العابدين الشرعية للبنات", "إناث", "حلب الجديدة"),
            ("الثانوية الشرعية للبنين في النيرب", "ذكور", "النيرب"),
            ("الثانوية الشرعية للبنات في النيرب", "إناث", "النيرب")
        ]
        c.executemany("INSERT INTO schools (name, gender, location) VALUES (?, ?, ?)", schools_data)
        
    conn.commit()
    conn.close()

init_db()

DOMAINS = ["التخطيط", "تنفيذ الدرس", "إدارة الصف", "شخصية المعلم", "المادة العلمية"]

CRITERIA = [
    {"domain": "التخطيط", "id": 1, "text": "يلتزم بتوزيع المنهاج", "max": 3},
    {"domain": "التخطيط", "id": 2, "text": "يخطط لتهيئة محفزة تراعي فيها المعلومات السابقة وتمهد للدرس الجديد", "max": 3},
    {"domain": "التخطيط", "id": 3, "text": "يصوغ الأهداف صياغة صحيحة للمجالات الثلاثة (المعرفي - المهاري - الوجداني)", "max": 4},
    {"domain": "التخطيط", "id": 4, "text": "يحدد استراتيجيات التدريس والأنشطة المناسبة لها", "max": 3},
    {"domain": "التخطيط", "id": 5, "text": "يحدد الوسائط المتعددة والبدائل التعليمية المناسبة", "max": 2},
    {"domain": "التخطيط", "id": 6, "text": "يخطط لتقويم لمراحله المختلفة (المرحلي - النهائي)", "max": 4},
    {"domain": "تنفيذ الدرس", "id": 7, "text": "يستثير دافعية المتعلمين نحو التعلم", "max": 4},
    {"domain": "تنفيذ الدرس", "id": 8, "text": "يوظف طرائق التدريس والأنشطة والوسائط بصورة مناسبة لمستويات المتعلمين وأهداف الدرس", "max": 6},
    {"domain": "تنفيذ الدرس", "id": 9, "text": "يستخدم السبورة بالشكل الأمثل للمحتوى التعليمي", "max": 4},
    {"domain": "تنفيذ الدرس", "id": 10, "text": "يراعي تنوع أهداف الدرس في التقويم وتنوع الأسئلة", "max": 4},
    {"domain": "تنفيذ الدرس", "id": 11, "text": "يعزز المتعلمين في الوقت المناسب وبالطريقة المناسبة", "max": 4},
    {"domain": "تنفيذ الدرس", "id": 12, "text": "يخصص الوقت المناسب لكل مرحلة من مراحل الدرس", "max": 3},
    {"domain": "تنفيذ الدرس", "id": 13, "text": "ينفذ ختاماً مناسباً للدرس", "max": 3},
    {"domain": "إدارة الصف", "id": 14, "text": "يهيئ بيئة صفية مادية ونفسية مريحة ومناسبة للمتعلمين", "max": 3},
    {"domain": "إدارة الصف", "id": 15, "text": "يوزع الاهتمام بين المتعلمين ويتعامل معهم بطريقة عادلة", "max": 3},
    {"domain": "إدارة الصف", "id": 16, "text": "يتبع قواعد صفية تنظم تفاعل وأنشطة المتعلمين بما يحقق أهداف الدرس", "max": 3},
    {"domain": "إدارة الصف", "id": 17, "text": "يراعي الفروق الفردية (فجوات الفاقد التعليمي) ويوزع الاهتمام بين المتعلمين", "max": 4},
    {"domain": "شخصية المعلم", "id": 18, "text": "يلتزم بالسمت الشرعي والمظهر اللائق", "max": 5},
    {"domain": "شخصية المعلم", "id": 19, "text": "يستخدم اللغة السليمة ويتنوع في نبرات الصوت ويستخدم لغة الجسد المناسبة", "max": 4},
    {"domain": "المادة العلمية", "id": 20, "text": "يهتم بالحيوية والنشاط داخل الصف", "max": 3},
    {"domain": "المادة العلمية", "id": 21, "text": "يركز على الجانب القيمي للطلاب ويستثمر الأمثلة لتعزيز القيم والأخلاق الحسنة", "max": 5},
    {"domain": "المادة العلمية", "id": 22, "text": "يتصف بالاتزان الانفعالي والذكاء العاطفي ويتقبل الملاحظات اللازمة للتطوير", "max": 3},
    {"domain": "المادة العلمية", "id": 23, "text": "يربط المادة العلمية بالحياة العملية ويحقق الترابط والتكامل مع المواد الأخرى", "max": 4},
    {"domain": "المادة العلمية", "id": 24, "text": "يراعي التدرج والتسلسل المنطقي للمادة والانتقال الصحيح", "max": 4},
    {"domain": "المادة العلمية", "id": 25, "text": "يظهر تمكناً من المادة العلمية التي يقدمها", "max": 12},
]

def calculate_domain_scores(scores_dict):
    dom_totals = {d: 0 for d in DOMAINS}
    for item in CRITERIA:
        actual = int(scores_dict.get(str(item['id']), item['max']))
        dom_totals[item['domain']] += actual
    return dom_totals

# -------------------------------------------------------------
# دالة HTML الرسمية
# -------------------------------------------------------------
def get_evaluation_html(eval_data):
    scores = json.loads(eval_data['scores_json']) if isinstance(eval_data['scores_json'], str) else eval_data['scores_json']
    
    rows_html = ""
    for item in CRITERIA:
        actual = scores.get(str(item['id']), item['max'])
        i_id = item['id']
        
        domain_td = ""
        if i_id == 1:
            domain_td = '<td rowspan="6" style="text-align: center; font-weight: bold; vertical-align: middle; background-color: #f9f9f9;">التخطيط</td>'
        elif i_id == 7:
            domain_td = '<td rowspan="7" style="text-align: center; font-weight: bold; vertical-align: middle; background-color: #f9f9f9;">تنفيذ الدرس</td>'
        elif i_id == 14:
            domain_td = '<td rowspan="4" style="text-align: center; font-weight: bold; vertical-align: middle; background-color: #f9f9f9;">إدارة الصف</td>'
        elif i_id == 18:
            domain_td = '<td rowspan="2" style="text-align: center; font-weight: bold; vertical-align: middle; background-color: #f9f9f9;">شخصية المعلم</td>'
        elif i_id == 20:
            domain_td = '<td rowspan="6" style="text-align: center; font-weight: bold; vertical-align: middle; background-color: #f9f9f9;">المادة العلمية</td>'

        notes_td = ""
        if i_id == 1:
            notes_td = f'<td rowspan="10" style="vertical-align: top; text-align: right; padding: 4px; font-size: 11px;"><b>نقاط التميز:</b><br>{eval_data.get("excellence_points", "")}</td>'
        elif i_id == 11:
            notes_td = f'<td rowspan="9" style="vertical-align: top; text-align: right; padding: 4px; font-size: 11px;"><b>نقاط التطوير:</b><br>{eval_data.get("dev_points", "")}</td>'
        elif i_id == 20:
            notes_td = f'<td rowspan="6" style="vertical-align: top; text-align: right; padding: 4px; font-size: 11px;"><b>المقترحات:</b><br>{eval_data.get("suggestions", "")}</td>'

        rows_html += f"""
        <tr>
            <td style="width: 5%; text-align: center; font-size: 12px;">{i_id}</td>
            {domain_td}
            <td style="width: 42%; text-align: right; font-size: 12px;">{item['text']}</td>
            <td style="width: 9%; text-align: center; font-size: 12px;">{item['max']}</td>
            <td style="width: 9%; text-align: center; font-weight: bold; font-size: 12px;">{actual}</td>
            {notes_td}
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="utf-8">
        <title>استمارة تقييم أداء المدرس</title>
        <style>
            @page {{
                size: A4;
                margin: 10mm;
            }}
            @media print {{
                body {{ margin: 0; padding: 0; }}
                .no-print {{ display: none !important; }}
            }}
            body {{
                font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
                direction: rtl;
                text-align: right;
                background-color: #fff;
                color: #000;
                padding: 5px;
            }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 4px; }}
            th, td {{ border: 1px solid #000; padding: 2.5px 4px; }}
            .bg-gray {{ background-color: #f2f2f2; }}
            .header-tbl td {{ border: none; font-weight: bold; }}
            .btn-print {{
                background-color: #0d5c3a; color: white; padding: 12px 18px;
                border: none; border-radius: 8px; cursor: pointer;
                font-size: 16px; font-weight: bold; margin-bottom: 12px; width: 100%;
            }}
        </style>
    </head>
    <body>
        <div class="no-print">
            <button class="btn-print" onclick="window.print()">🖨️ اضغط هنا للطباعة أو الحفظ كـ PDF عبر المتصفح</button>
        </div>
        <table class="header-tbl">
            <tr>
                <td style="width: 30%; text-align: right; font-size: 13px;">مديرية أوقاف حلب</td>
                <td style="width: 40%; text-align: center; font-size: 16px;">لجنة تقييم أداء المدرس</td>
                <td style="width: 30%; text-align: left; font-size: 13px;">رقم اللجنة: {eval_data['committee_no']}</td>
            </tr>
        </table>
        <table>
            <tr>
                <td class="bg-gray" style="width: 15%; font-weight: bold;">التاريخ:</td>
                <td style="width: 35%;">{eval_data['visit_date']}</td>
                <td class="bg-gray" style="width: 15%; font-weight: bold;">المؤسسة التعليمية الشرعية:</td>
                <td style="width: 35%;">{eval_data['school_name']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">العام الدراسي:</td>
                <td>{eval_data['academic_year']}</td>
                <td class="bg-gray" style="font-weight: bold;">الفصل الدراسي:</td>
                <td>{eval_data['semester']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">المدرس:</td>
                <td style="font-weight: bold;">{eval_data['teacher_name']}</td>
                <td class="bg-gray" style="font-weight: bold;">المسؤول العلمي:</td>
                <td>{eval_data['supervisor_name']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">المادة:</td>
                <td>{eval_data['subject']}</td>
                <td class="bg-gray" style="font-weight: bold;">الاختصاص:</td>
                <td>{eval_data['specialization']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">عدد الطلاب:</td>
                <td>{eval_data['student_count']}</td>
                <td class="bg-gray" style="font-weight: bold;">الصف:</td>
                <td>{eval_data['grade_level']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">الشعبة:</td>
                <td>{eval_data['section']}</td>
                <td class="bg-gray" style="font-weight: bold;">الموضوع:</td>
                <td>{eval_data['lesson_topic']}</td>
            </tr>
            <tr>
                <td class="bg-gray" style="font-weight: bold;">الخبرة في التعليم:</td>
                <td>{eval_data['experience']}</td>
                <td class="bg-gray" style="font-weight: bold;">الوضع الوظيفي:</td>
                <td>{eval_data['job_status']}</td>
            </tr>
        </table>
        <table>
            <thead>
                <tr class="bg-gray" style="text-align: center; font-weight: bold; font-size: 11.5px;">
                    <th style="width: 5%;">م</th>
                    <th style="width: 14%;">مجال التقييم</th>
                    <th style="width: 42%;">بنود اللائحة</th>
                    <th style="width: 9%;">الدرجة المستحقة</th>
                    <th style="width: 9%;">الدرجة الفعلية</th>
                    <th style="width: 21%;">الملاحظات</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        <table>
            <tr class="bg-gray" style="font-weight: bold; text-align: center; font-size: 12px;">
                <td style="width: 50%;">مجموع الدرجات: {eval_data['total_score']} / 100</td>
                <td style="width: 50%;">التقدير النهائي: {eval_data['rating']}</td>
            </tr>
            <tr>
                <td colspan="2" style="text-align: center; font-size: 10px; padding: 2px;">
                    ممتاز: 90–100 | جيد جداً: 80–89 | جيد: 70–79 | مقبول: 50–69 | ضعيف: أقل من 50
                </td>
            </tr>
        </table>
        <table style="border: none; margin-top: 12px;">
            <tr style="font-weight: bold; text-align: center;">
                <td style="border: none; width: 25%;">المدرسة</td>
                <td style="border: none; width: 25%;">الموجّه الاختصاصي</td>
                <td style="border: none; width: 25%;">شعبة التوجيه الاختصاصي</td>
                <td style="border: none; width: 25%;">رئيس إدارة التعليم الشرعي</td>
            </tr>
            <tr style="text-align: center;">
                <td style="border: none; padding-top: 22px;">.....................</td>
                <td style="border: none; padding-top: 22px;">{eval_data['supervisor_name']}</td>
                <td style="border: none; padding-top: 22px;">.....................</td>
                <td style="border: none; padding-top: 22px;">.....................</td>
            </tr>
        </table>
    </body>
    </html>
    """
    return html

def generate_direct_pdf_bytes(eval_data):
    if not WEASYPRINT_AVAILABLE:
        return None
    try:
        html_str = get_evaluation_html(eval_data)
        pdf_bytes = HTML(string=html_str).write_pdf()
        return pdf_bytes
    except Exception:
        return None

# -------------------------------------------------------------
# دالة توليد استمارة Excel الرسمية
# -------------------------------------------------------------
def generate_evaluation_excel_form(eval_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "استمارة التقييم"
    ws.views.sheetView[0].rightToLeft = True

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_norm = Font(name="Arial", size=9)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.merge_cells("A1:B1")
    ws["A1"] = "مديرية أوقاف حلب"
    ws["A1"].font = font_bold
    ws["A1"].alignment = align_right

    ws.merge_cells("C1:D1")
    ws["C1"] = "لجنة تقييم أداء المدرس"
    ws["C1"].font = Font(name="Arial", size=13, bold=True)
    ws["C1"].alignment = align_center

    ws.merge_cells("E1:F1")
    ws["E1"] = f"رقم اللجنة: {eval_data['committee_no']}"
    ws["E1"].font = font_bold
    ws["E1"].alignment = Alignment(horizontal="left", vertical="center")

    meta_rows = [
        ("التاريخ:", str(eval_data['visit_date']), "المؤسسة التعليمية الشرعية:", eval_data['school_name']),
        ("العام الدراسي:", eval_data['academic_year'], "الفصل الدراسي:", eval_data['semester']),
        ("المدرس:", eval_data['teacher_name'], "المسؤول العلمي:", eval_data['supervisor_name']),
        ("المادة:", eval_data['subject'], "الاختصاص:", eval_data['specialization']),
        ("عدد الطلاب:", eval_data['student_count'], "الصف:", eval_data['grade_level']),
        ("الشعبة:", eval_data['section'], "الموضوع:", eval_data['lesson_topic']),
        ("الخبرة في التعليم:", eval_data['experience'], "الوضع الوظيفي:", eval_data['job_status'])
    ]

    r_idx = 3
    for r in meta_rows:
        ws.cell(row=r_idx, column=1, value=r[0]).fill = fill_gray
        ws.cell(row=r_idx, column=1).font = font_bold
        ws.cell(row=r_idx, column=2, value=r[1]).font = font_norm
        ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=3)
        
        ws.cell(row=r_idx, column=4, value=r[2]).fill = fill_gray
        ws.cell(row=r_idx, column=4).font = font_bold
        ws.cell(row=r_idx, column=5, value=r[3]).font = font_norm
        ws.merge_cells(start_row=r_idx, start_column=5, end_row=r_idx, end_column=6)
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = border
            ws.cell(row=r_idx, column=c).alignment = align_center
        r_idx += 1

    r_idx += 1
    headers = ["م", "مجال التقييم", "بنود اللائحة", "الدرجة المستحقة", "الدرجة الفعلية", "الملاحظات"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=r_idx, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_gray
        cell.alignment = align_center
        cell.border = border

    scores = json.loads(eval_data['scores_json']) if isinstance(eval_data['scores_json'], str) else eval_data['scores_json']
    start_eval_row = r_idx + 1

    for item in CRITERIA:
        r_idx += 1
        actual = scores.get(str(item['id']), item['max'])
        ws.cell(row=r_idx, column=1, value=item['id']).alignment = align_center
        ws.cell(row=r_idx, column=2, value=item['domain']).alignment = align_center
        ws.cell(row=r_idx, column=3, value=item['text']).alignment = align_right
        ws.cell(row=r_idx, column=4, value=item['max']).alignment = align_center
        ws.cell(row=r_idx, column=5, value=actual).alignment = align_center
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = border
            ws.cell(row=r_idx, column=c).font = font_norm

    ws.merge_cells(f"B{start_eval_row}:B{start_eval_row+5}")
    ws.merge_cells(f"B{start_eval_row+6}:B{start_eval_row+12}")
    ws.merge_cells(f"B{start_eval_row+13}:B{start_eval_row+16}")
    ws.merge_cells(f"B{start_eval_row+17}:B{start_eval_row+18}")
    ws.merge_cells(f"B{start_eval_row+19}:B{r_idx}")

    ws.merge_cells(f"F{start_eval_row}:F{start_eval_row+9}")
    ws[f"F{start_eval_row}"] = f"نقاط التميز:\n{eval_data.get('excellence_points', '')}"
    ws[f"F{start_eval_row}"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    ws.merge_cells(f"F{start_eval_row+10}:F{start_eval_row+18}")
    ws[f"F{start_eval_row+10}"] = f"نقاط التطوير:\n{eval_data.get('dev_points', '')}"
    ws[f"F{start_eval_row+10}"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    ws.merge_cells(f"F{start_eval_row+19}:F{r_idx}")
    ws[f"F{start_eval_row+19}"] = f"المقترحات:\n{eval_data.get('suggestions', '')}"
    ws[f"F{start_eval_row+19}"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    r_idx += 1
    ws.merge_cells(f"A{r_idx}:C{r_idx}")
    ws[f"A{r_idx}"] = f"مجموع الدرجات: {eval_data['total_score']} / 100"
    ws[f"A{r_idx}"].font = font_bold
    ws[f"A{r_idx}"].alignment = align_center
    ws[f"A{r_idx}"].fill = fill_gray

    ws.merge_cells(f"D{r_idx}:F{r_idx}")
    ws[f"D{r_idx}"] = f"التقدير النهائي: {eval_data['rating']}"
    ws[f"D{r_idx}"].font = font_bold
    ws[f"D{r_idx}"].alignment = align_center
    ws[f"D{r_idx}"].fill = fill_gray

    for c in range(1, 7):
        ws.cell(row=r_idx, column=c).border = border

    r_idx += 2
    sigs = ["المدرسة", "", "الموجّه الاختصاصي", "", "شعبة التوجيه الاختصاصي", "رئيس إدارة التعليم الشرعي"]
    for i, s in enumerate(sigs, 1):
        ws.cell(row=r_idx, column=i, value=s).font = font_bold
        ws.cell(row=r_idx, column=i).alignment = align_center

    col_widths = {1: 5, 2: 15, 3: 45, 4: 12, 5: 12, 6: 28}
    for c, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

# -------------------------------------------------------------
# دالة توليد التقرير التركيبي
# -------------------------------------------------------------
def generate_annual_executive_report(evals_df, schools_df, users_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        total_schools = len(schools_df)
        visited_schools = len(evals_df['school_name'].unique()) if not evals_df.empty else 0
        coverage_pct = (visited_schools / total_schools * 100) if total_schools > 0 else 0
        
        rating_counts = eval_data_rating = evals_df['rating'].value_counts() if not evals_df.empty else pd.Series()
        ratings_data = []
        for r in ["ممتاز", "جيد جداً", "جيد", "مقبول", "ضعيف"]:
            cnt = rating_counts.get(r, 0)
            pct = (cnt / len(evals_df) * 100) if len(evals_df) > 0 else 0
            ratings_data.append({"التقدير": r, "العدد": cnt, "النسبة المئوية": f"{pct:.1f}%"})
            
        summary_kpi_df = pd.DataFrame([
            {"المؤشر": "إجمالي عدد المدارس الشرعية", "القيمة": total_schools},
            {"المؤشر": "المدارس التي تمت زيارتها", "القيمة": visited_schools},
            {"المؤشر": "نسبة التغطية الميدانية", "القيمة": f"{coverage_pct:.1f}%"},
            {"المؤشر": "إجمالي الزيارات والتقييمات المنجزة", "القيمة": len(evals_df)},
            {"المؤشر": "متوسط درجات التقييم العام", "القيمة": f"{evals_df['total_score'].mean():.1f} / 100" if not evals_df.empty else "0"}
        ])
        
        summary_kpi_df.to_excel(writer, sheet_name='المؤشرات العامة ونسب التقديرات', index=False, startrow=0)
        pd.DataFrame(ratings_data).to_excel(writer, sheet_name='المؤشرات العامة ونسب التقديرات', index=False, startrow=8)
        
        if not evals_df.empty:
            subj_perf = evals_df.groupby('subject').agg(
                عدد_التقييمات=('id', 'count'),
                متوسط_الدرجة=('total_score', 'mean'),
                أعلى_درجة=('total_score', 'max'),
                أدنى_درجة=('total_score', 'min')
            ).reset_index().rename(columns={'subject': 'المادة الدراسية'})
            subj_perf['متوسط_الدرجة'] = subj_perf['متوسط_الدرجة'].round(1)
            subj_perf.sort_values(by='متوسط_الدرجة', ascending=False).to_excel(writer, sheet_name='مقارنة أداء المواد الدراسية', index=False)
            
        if not evals_df.empty:
            evals_df[['teacher_name', 'school_name', 'subject', 'visit_date', 'total_score', 'rating', 'supervisor_name']].to_excel(
                writer, sheet_name='سجل التقييمات الشامل', index=False
            )

    output.seek(0)
    wb = openpyxl.load_workbook(output)
    for ws in wb.worksheets:
        ws.views.sheetView[0].rightToLeft = True
    excel_final = io.BytesIO()
    wb.save(excel_final)
    excel_final.seek(0)
    return excel_final.getvalue()

# -------------------------------------------------------------
# تسجيل الدخول
# -------------------------------------------------------------
if "user" not in st.session_state:
    st.session_state.user = None

def login(username, password):
    conn = sqlite3.connect("evaluation_system.db")
    c = conn.cursor()
    c.execute("SELECT id, username, full_name, specialization, role FROM users WHERE username=? AND password=?", (username, password))
    user = c.fetchone()
    conn.close()
    return user

if st.session_state.user is None:
    # تنسيق شاشة تسجيل الدخول في المنتصف بحجم متناسق وأنيق
    st.markdown("<div style='height: 40px;'></div>", unsafe_allow_html=True)
    col_l, col_center, col_r = st.columns([1, 2, 1])
    
    with col_center:
        st.markdown("""
        <div style='text-align: center; margin-bottom: 20px;'>
            <h2 style='font-size: 28px; font-weight: 800; color: #0d5c3a; margin-bottom: 4px;'>🕌 منظومة التقييم الشرعي</h2>
            <p style='font-size: 15px; color: #64748b;'>تسجيل دخول الموجهين والمشرفين</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div class='mobile-card'>", unsafe_allow_html=True)
        u_name = st.text_input("اسم المستخدم", placeholder="اسم المستخدم...")
        u_pass = st.text_input("كلمة المرور", type="password", placeholder="••••••")
        
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        if st.button("🔑 تسجيل الدخول", type="primary", use_container_width=True):
            user_data = login(u_name, u_pass)
            if user_data:
                st.session_state.user = {
                    "id": user_data[0], "username": user_data[1],
                    "name": user_data[2], "specialization": user_data[3], "role": user_data[4]
                }
                st.rerun()
            else:
                st.error("اسم المستخدم أو كلمة المرور غير صحيحة")
        st.markdown("</div>", unsafe_allow_html=True)
        
    st.stop()

# -------------------------------------------------------------
# الرأس وشريط التنقل
# -------------------------------------------------------------
st.markdown(f"""
<div style='display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; background: #fff; padding: 12px 16px; border-radius: 14px; box-shadow: 0 3px 8px rgba(0,0,0,0.05);'>
    <div>
        <span style='font-size: 18px; font-weight: 800; color: #0d5c3a;'>👤 {st.session_state.user['name']}</span><br>
        <span style='font-size: 13px; color: #64748b;'>{st.session_state.user['specialization']} | {st.session_state.user['role']}</span>
    </div>
</div>
""", unsafe_allow_html=True)

main_menu_options = ["📝 استمارة تقييم جديدة", "🔍 سجل الزيارات والتصدير", "📈 تتبع تطور المدرسين"]
if st.session_state.user["role"] == "Admin":
    main_menu_options = [
        "📊 لوحة المؤشرات والمتابعة",
        "📑 التقرير التركيبي السنوي",
        "📝 استمارة تقييم جديدة",
        "🔍 سجل الزيارات والتصدير",
        "📈 تتبع تطور المدرسين",
        "⚙️ إدارة النظام والحماية"
    ]

c_nav, c_out = st.columns([4, 1])
with c_nav:
    choice = st.radio("القائمة الرئيسية", main_menu_options, horizontal=True, label_visibility="collapsed")
with c_out:
    if st.button("🚪 خروج", use_container_width=True):
        st.session_state.user = None
        st.rerun()

# -------------------------------------------------------------
# 0. لوحة المؤشرات (Admin)
# -------------------------------------------------------------
if choice == "📊 لوحة المؤشرات والمتابعة" and st.session_state.user["role"] == "Admin":
    st.markdown("### 📊 لوحة المتابعة الميدانية وخريطة الزيارات")
    
    conn = sqlite3.connect("evaluation_system.db")
    schools_df = pd.read_sql_query("SELECT * FROM schools", conn)
    evals_df = pd.read_sql_query("SELECT * FROM evaluations", conn)
    users_df = pd.read_sql_query("SELECT * FROM users WHERE role='Supervisor'", conn)
    conn.close()
    
    total_schools_count = len(schools_df)
    visited_schools_set = set(evals_df["school_name"].unique()) if not evals_df.empty else set()
    visited_count = len(visited_schools_set)
    unvisited_count = total_schools_count - visited_count
    total_visits = len(evals_df)
    
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.markdown(f"<div class='mobile-card' style='text-align: center;'><div style='color:#64748b; font-size:13px;'>إجمالي المدارس</div><div style='font-size:26px; font-weight:900;'>{total_schools_count}</div></div>", unsafe_allow_html=True)
    with m2:
        st.markdown(f"<div class='mobile-card' style='text-align: center;'><div style='color:#166534; font-size:13px;'>مدارس زِيرَت</div><div style='font-size:26px; font-weight:900; color:#166534;'>{visited_count}</div></div>", unsafe_allow_html=True)
    with m3:
        st.markdown(f"<div class='mobile-card' style='text-align: center;'><div style='color:#991b1b; font-size:13px;'>مدارس لم تُزَر</div><div style='font-size:26px; font-weight:900; color:#991b1b;'>{unvisited_count}</div></div>", unsafe_allow_html=True)
    with m4:
        st.markdown(f"<div class='mobile-card' style='text-align: center;'><div style='color:#0d5c3a; font-size:13px;'>إجمالي الزيارات</div><div style='font-size:26px; font-weight:900; color:#0d5c3a;'>{total_visits}</div></div>", unsafe_allow_html=True)

    tab_vis, tab_unvis, tab_subjects, tab_monthly, tab_sups = st.tabs([
        "🏫 المدارس وزمن الانقطاع",
        "🚨 مدارس لم تُزَر قط",
        "📚 مقارنة جودة المواد",
        "📅 الزيارات الشهرية",
        "👥 نشاط الموجهين"
    ])

    with tab_vis:
        school_status_data = []
        today_date = date.today()
        for _, sc in schools_df.iterrows():
            s_name = sc['name']
            sc_evals = evals_df[evals_df['school_name'] == s_name]
            if not sc_evals.empty:
                sc_evals_copy = sc_evals.copy()
                sc_evals_copy['v_date'] = pd.to_datetime(sc_evals_copy['visit_date'], errors='coerce')
                latest_date = sc_evals_copy['v_date'].max()
                days_diff = (pd.Timestamp(today_date) - latest_date).days if pd.notnull(latest_date) else 0
                status_desc = f"منذ {days_diff} يوم"
                if days_diff > 30: status_desc = f"⚠️ تنبيه انقطاع ({days_diff} يوم)"
                if days_diff > 60: status_desc = f"🚨 انقطاع حرج ({days_diff} يوم)"
                school_status_data.append({
                    "المدرسة": s_name, "النوع": sc['gender'], "الموقع": sc['location'],
                    "الزيارات الكلية": len(sc_evals), "آخر زيارة": latest_date.strftime('%Y-%m-%d') if pd.notnull(latest_date) else "—",
                    "حالة الانقطاع": status_desc
                })
        status_df = pd.DataFrame(school_status_data)
        if not status_df.empty:
            st.dataframe(status_df, use_container_width=True)
        else:
            st.info("لا توجد زيارات مسجلة حتى الآن.")

    with tab_unvis:
        unvisited_schools = schools_df[~schools_df['name'].isin(visited_schools_set)]
        if unvisited_schools.empty:
            st.success("🎉 رائع! تم تغطية وزيارة جميع المدارس.")
        else:
            st.warning(f"⚠️ يوجد ({len(unvisited_schools)}) مدرسة لم تتم زيارتها حتى الآن:")
            st.dataframe(unvisited_schools[["name", "gender", "location"]].rename(columns={"name": "المدرسة", "gender": "النوع", "location": "الموقع"}), use_container_width=True)

    with tab_subjects:
        st.markdown("#### 📚 مقارنة متوسط الدرجات ومؤشر الأداء بين المواد:")
        if not evals_df.empty:
            subj_stats = evals_df.groupby('subject').agg(
                عدد_الزيارات=('id', 'count'),
                متوسط_الدرجة=('total_score', 'mean'),
                أعلى_درجة=('total_score', 'max'),
                أدنى_درجة=('total_score', 'min')
            ).reset_index().rename(columns={'subject': 'المادة الدراسية'})
            subj_stats['متوسط_الدرجة'] = subj_stats['متوسط_الدرجة'].round(1)
            subj_stats = subj_stats.sort_values(by='متوسط_الدرجة', ascending=False)
            st.dataframe(subj_stats, use_container_width=True)
            chart_subj = subj_stats[['المادة الدراسية', 'متوسط_الدرجة']].set_index('المادة الدراسية')
            st.bar_chart(chart_subj)
        else:
            st.info("لا توجد بيانات كافية لعرض المقارنة.")

    with tab_monthly:
        if not evals_df.empty:
            evals_df['v_dt'] = pd.to_datetime(evals_df['visit_date'], errors='coerce')
            evals_df['الشهر-السنة'] = evals_df['v_dt'].dt.strftime('%m-%Y')
            monthly_pivot = pd.pivot_table(evals_df, values='id', index='school_name', columns='الشهر-السنة', aggfunc='count', fill_value=0)
            monthly_pivot.index.name = "اسم المدرسة"
            st.dataframe(monthly_pivot, use_container_width=True)
        else:
            st.info("لا توجد بيانات متاحة.")

    with tab_sups:
        if not evals_df.empty:
            sup_counts = evals_df['supervisor_name'].value_counts().reset_index()
            sup_counts.columns = ['اسم الموجه', 'عدد الزيارات المنجزة']
            all_sups = users_df[['full_name', 'specialization']].rename(columns={'full_name': 'اسم الموجه', 'specialization': 'التخصص'})
            merged_sups = pd.merge(all_sups, sup_counts, on='اسم الموجه', how='left').fillna(0)
            merged_sups['عدد الزيارات المنجزة'] = merged_sups['عدد الزيارات المنجزة'].astype(int)
            st.dataframe(merged_sups.sort_values(by='عدد الزيارات المنجزة', ascending=False), use_container_width=True)

# -------------------------------------------------------------
# 1. التقرير التركيبي السنوي (Admin)
# -------------------------------------------------------------
elif choice == "📑 التقرير التركيبي السنوي" and st.session_state.user["role"] == "Admin":
    st.markdown("### 📑 التقرير التركيبي الفصلي والسنوي المجمع")
    st.caption("تقرير رسمي شامل يلخص إنجاز ونسب التقييم والتغطية الميدانية لرفعه إلى رئيس إدارة التعليم الشرعي.")
    
    conn = sqlite3.connect("evaluation_system.db")
    evals_df = pd.read_sql_query("SELECT * FROM evaluations", conn)
    schools_df = pd.read_sql_query("SELECT * FROM schools", conn)
    users_df = pd.read_sql_query("SELECT * FROM users", conn)
    conn.close()
    
    if evals_df.empty:
        st.info("لا توجد تقييمات مسجلة بعد لإعداد التقرير التركيبي.")
    else:
        tot_evals = len(evals_df)
        tot_sch = len(schools_df)
        vis_sch = len(evals_df['school_name'].unique())
        cov_pct = (vis_sch / tot_sch * 100) if tot_sch > 0 else 0
        avg_score = evals_df['total_score'].mean()
        
        st.markdown(f"""
        <div class='mobile-card'>
            <h4>🏛️ خلاصة إنجاز التعليم الشرعي</h4>
            • نسبة التغطية الميدانية للمدارس: <b>{cov_pct:.1f}%</b> ({vis_sch} من أصل {tot_sch} مدرسة)<br>
            • إجمالي الزيارات المنفذة: <b>{tot_evals}</b> زيارة تفتيشية وتوجيهية<br>
            • متوسط الأداء العام للمدرسين: <b>{avg_score:.1f} / 100</b>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("#### 🎯 توزيع التقديرات العامة:")
        r_counts = evals_df['rating'].value_counts()
        r_data = []
        for r in ["ممتاز", "جيد جداً", "جيد", "مقبول", "ضعيف"]:
            c_val = r_counts.get(r, 0)
            p_val = (c_val / tot_evals * 100) if tot_evals > 0 else 0
            r_data.append({"التقدير": r, "العدد": c_val, "النسبة المئوية": f"{p_val:.1f}%"})
        st.dataframe(pd.DataFrame(r_data), use_container_width=True)
        
        report_excel_bytes = generate_annual_executive_report(evals_df, schools_df, users_df)
        st.download_button(
            label="📊 تحميل التقرير التركيبي السنوي الرسمي (Excel)",
            data=report_excel_bytes,
            file_name=f"التقرير_التركيبي_السنوي_للتعليم_الشرعي_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

# -------------------------------------------------------------
# 2. شاشة استمارة تقييم جديدة
# -------------------------------------------------------------
elif choice == "📝 استمارة تقييم جديدة":
    
    if "last_saved_eval" in st.session_state and st.session_state.last_saved_eval:
        last = st.session_state.last_saved_eval
        st.markdown(f"""
        <div class='success-box'>
            <div style='font-size: 38px;'>✅</div>
            <div style='font-size: 22px; font-weight: 900; margin-bottom: 6px;'>تم حفظ واعتماد الاستمارة بنجاح!</div>
            <div style='font-size: 16px;'>تم تسجيل تقييم المدرس: <b>{last['teacher']}</b></div>
            <div style='font-size: 14px; opacity: 0.9;'>المؤسسة: {last['school']} | الدرجة: {last['score']}/100 ({last['rating']})</div>
            <div style='font-size: 13px; margin-top: 8px; background: rgba(255,255,255,0.2); padding: 4px 12px; border-radius: 10px; display: inline-block;'>
                ⚡️ جاري رفع الشواهد والإكسل إلى Google Drive في الخلفية
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("➕ تعبئة استمارة جديدة أخرى", type="secondary", use_container_width=True):
            st.session_state.last_saved_eval = None
            st.rerun()

    for item in CRITERIA:
        k = f"q_val_{item['id']}"
        if k not in st.session_state:
            st.session_state[k] = item['max']

    total_live = sum(st.session_state[f"q_val_{item['id']}"] for item in CRITERIA)
    rating_live = "ممتاز" if total_live >= 90 else "جيد جداً" if total_live >= 80 else "جيد" if total_live >= 70 else "مقبول" if total_live >= 50 else "ضعيف"

    st.markdown(f"""
    <div class='score-banner'>
        <div style='font-size: 14px; opacity: 0.9;'>المجموع الكلي المباشر</div>
        <div style='font-size: 34px; font-weight: 900; margin: 4px 0;'>{total_live} <span style='font-size: 16px; font-weight: 600;'>/ 100</span></div>
        <div style='display: inline-block; background: rgba(255,255,255,0.25); padding: 4px 16px; border-radius: 20px; font-size: 14px; font-weight: 800;'>
            التقدير: {rating_live}
        </div>
    </div>
    """, unsafe_allow_html=True)

    tab_info, tab_domains, tab_notes = st.tabs(["📌 1. بيانات الزيارة", "📋 2. مجالات التقييم", "✍️ 3. الملاحظات والرفع"])

    with tab_info:
        st.markdown("<div class='mobile-card'>", unsafe_allow_html=True)
        
        st.markdown("#### 🏫 تصنيف ونوع المدرسة:")
        gender_type = st.radio(
            "اختر نوع المدرسة لتصفية القائمة:",
            ["ذكور", "إناث"],
            horizontal=True,
            index=0,
            key="school_gender_radio"
        )
        
        conn = sqlite3.connect("evaluation_system.db")
        filtered_schools_df = pd.read_sql_query("SELECT name FROM schools WHERE gender=?", conn, params=(gender_type,))
        sups_all = pd.read_sql_query("SELECT full_name FROM users WHERE role='Supervisor'", conn)["full_name"].tolist()
        conn.close()
        
        schools_list = filtered_schools_df['name'].tolist() if not filtered_schools_df.empty else [f"لا توجد مدارس {gender_type}"]

        c1, c2 = st.columns(2)
        with c1:
            school_name = st.selectbox(f"المؤسسة التعليمية الشرعية ({'بنين' if gender_type == 'ذكور' else 'بنات'})", schools_list)
            teacher_name = st.text_input("اسم المدرس *", placeholder="أدخل اسم المدرس...")
            subject = st.text_input("المادة", value=st.session_state.user["specialization"] if st.session_state.user["role"] != "Admin" else "شريعة")
            grade_level = st.selectbox("الصف", ["السابع", "الثامن", "التاسع", "العاشر", "الحادي عشر", "الثالث الثانوي"])
            section = st.text_input("الشعبة", value="الأولى")
        with c2:
            visit_date = st.date_input("تاريخ الزيارة", value=date.today())
            academic_year = st.selectbox("العام الدراسي", ["2026 - 2027", "2025 - 2026"])
            semester = st.selectbox("الفصل الدراسي", ["الفصل الأول", "الفصل الثاني"])
            committee_no = st.text_input("رقم اللجنة", value="1")

        if st.session_state.user["role"] == "Admin":
            supervisor_name = st.selectbox("المسؤول العلمي / الموجه المشرف", sups_all if sups_all else [st.session_state.user["name"]])
        else:
            supervisor_name = st.session_state.user["name"]

        lesson_topic = st.text_input("موضوع الدرس", placeholder="اكتب موضوع الدرس...")
        
        c3, c4 = st.columns(2)
        with c3:
            specialization = st.text_input("الاختصاص", value=subject)
            student_count = st.number_input("عدد الطلاب", min_value=1, value=25)
        with c4:
            job_status = st.selectbox("الوضع الوظيفي", ["أصيل", "وكيل", "مكلف"])
            experience = st.text_input("الخبرة في التعليم", value="5 سنوات")
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_domains:
        domain_tabs = st.tabs(DOMAINS)
        for idx, dom in enumerate(DOMAINS):
            with domain_tabs[idx]:
                dom_items = [i for i in CRITERIA if i['domain'] == dom]
                for it in dom_items:
                    st.markdown(f"""
                    <div class='criterion-item'>
                        <div style='font-weight: 800; font-size: 15px; color: #1e293b;'>{it['id']}. {it['text']}</div>
                        <div style='font-size: 13px; color: #64748b; margin-bottom: 4px;'>الدرجة القصوى المستحقة: ({it['max']} درجات)</div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    scores_options = list(range(it['max'], -1, -1))
                    
                    st.selectbox(
                        f"اختر درجة البند {it['id']}:",
                        options=scores_options,
                        index=scores_options.index(st.session_state[f"q_val_{it['id']}"]) if st.session_state[f"q_val_{it['id']}"] in scores_options else 0,
                        key=f"select_val_{it['id']}",
                        format_func=lambda x, m=it['max']: f"⭐️ {x} من {m}" if x == m else f"⚪️ {x} من {m}",
                        on_change=lambda i=it['id']: st.session_state.update({f"q_val_{i}": st.session_state[f"select_val_{i}"]})
                    )
                    st.markdown("<div style='height: 6px;'></div>", unsafe_allow_html=True)

    with tab_notes:
        st.markdown("<div class='mobile-card'>", unsafe_allow_html=True)
        excellence_points = st.text_area("🌟 نقاط التميز", placeholder="اكتب نقاط القوة والتميز...")
        dev_points = st.text_area("💡 نقاط التطوير", placeholder="اكتب نقاط التطوير والتحسين...")
        suggestions = st.text_area("📌 المقترحات والتوصيات", placeholder="اكتب المقترحات والتوجيهات...")
        uploaded_files = st.file_uploader("📷 رفع شواهد وصور / فيديو من الهاتف", accept_multiple_files=True)
        st.markdown("</div>", unsafe_allow_html=True)

        if st.button("💾 حفظ واعتماد الاستمارة فوراً", type="primary", use_container_width=True):
            if not teacher_name.strip():
                st.error("⚠️ يرجى إدخال اسم المدرس في تبويب البيانات الأساسية.")
            else:
                final_scores_dict = {str(item['id']): st.session_state[f"q_val_{item['id']}"] for item in CRITERIA}
                
                saved_media = []
                uploaded_files_data = []
                if uploaded_files:
                    os.makedirs("uploads", exist_ok=True)
                    for f in uploaded_files:
                        f_bytes = f.getbuffer()
                        path = os.path.join("uploads", f.name)
                        with open(path, "wb") as file_out:
                            file_out.write(f_bytes)
                        saved_media.append(path)
                        uploaded_files_data.append((f.name, f.getvalue(), f.type))

                eval_data_dict = {
                    "committee_no": committee_no, "visit_date": str(visit_date), "academic_year": academic_year,
                    "semester": semester, "school_name": school_name, "gender_type": gender_type,
                    "supervisor_name": supervisor_name, "teacher_name": teacher_name, "subject": subject,
                    "specialization": specialization, "student_count": student_count, "grade_level": grade_level,
                    "section": section, "lesson_topic": lesson_topic, "job_status": job_status,
                    "experience": experience, "scores_json": json.dumps(final_scores_dict), "total_score": total_live,
                    "rating": rating_live, "excellence_points": excellence_points, "dev_points": dev_points,
                    "suggestions": suggestions
                }

                conn = sqlite3.connect("evaluation_system.db")
                c = conn.cursor()
                c.execute('''INSERT INTO evaluations (
                    committee_no, visit_date, academic_year, semester, school_name, gender_type,
                    supervisor_name, teacher_name, subject, specialization, student_count,
                    grade_level, section, lesson_topic, job_status, experience,
                    scores_json, total_score, rating, excellence_points, dev_points, suggestions,
                    media_paths, drive_links, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                    committee_no, str(visit_date), academic_year, semester, school_name, gender_type,
                    supervisor_name, teacher_name, subject, specialization, student_count,
                    grade_level, section, lesson_topic, job_status, experience,
                    json.dumps(final_scores_dict), total_live, rating_live, excellence_points, dev_points, suggestions,
                    ",".join(saved_media), "", "معتمد"
                ))
                new_eval_id = c.lastrowid
                conn.commit()
                conn.close()

                threading.Thread(
                    target=background_upload_task,
                    args=(new_eval_id, eval_data_dict, uploaded_files_data, school_name, visit_date)
                ).start()

                st.session_state.last_saved_eval = {
                    "teacher": teacher_name,
                    "school": school_name,
                    "score": total_live,
                    "rating": rating_live
                }
                st.rerun()

# -------------------------------------------------------------
# 3. سجل الزيارات وتصفية المدارس بحسب النوع
# -------------------------------------------------------------
elif choice == "🔍 سجل الزيارات والتصدير":
    st.markdown("### 🔍 سجل الزيارات واستمارات التقييم")
    
    conn = sqlite3.connect("evaluation_system.db")
    if st.session_state.user["role"] == "Admin":
        df = pd.read_sql_query("SELECT * FROM evaluations ORDER BY id DESC", conn)
    else:
        df = pd.read_sql_query("SELECT * FROM evaluations WHERE supervisor_name=? ORDER BY id DESC", conn, params=(st.session_state.user["name"],))
    conn.close()
    
    if df.empty:
        st.info("لا توجد استمارات مسجلة حتى الآن.")
    else:
        with st.expander("🗑️ قسم حذف استمارة محددة فوراً", expanded=False):
            delete_candidates = df.apply(lambda r: f"ID: {r['id']} | المدرس: {r['teacher_name']} | المدرسة: {r['school_name']} | التاريخ: {r['visit_date']}", axis=1).tolist()
            del_selection = st.selectbox("اختر الاستمارة المطلوب حذفها:", delete_candidates, key="del_select_box")
            if st.button("🗑️ تأكيد حذف هذه الاستمارة الآن", type="secondary", use_container_width=True):
                selected_del_id = int(del_selection.split("|")[0].replace("ID:", "").strip())
                delete_evaluation_by_id(selected_del_id)
                st.success(f"✅ تم حذف الاستمارة رقم {selected_del_id} بنجاح!")
                st.rerun()

        st.markdown("---")
        
        c_flt1, c_flt2 = st.columns([2, 2])
        with c_flt1:
            gender_filter = st.radio("تصفية بحسب نوع المدرسة:", ["الكل", "ذكور", "إناث"], horizontal=True)
        with c_flt2:
            search_txt = st.text_input("🔍 بحث باسم المدرس أو المدرسة", placeholder="اكتب للبحث...")

        filtered_df = df.copy()
        if gender_filter != "الكل":
            filtered_df = filtered_df[filtered_df["gender_type"] == gender_filter]
        if search_txt:
            filtered_df = filtered_df[
                filtered_df["teacher_name"].str.contains(search_txt, na=False) |
                filtered_df["school_name"].str.contains(search_txt, na=False)
            ]

        for _, row in filtered_df.iterrows():
            eval_id = row['id']
            with st.expander(f"📄 {row['teacher_name']} — {row['school_name']} ({row['total_score']} / 100 - {row['rating']})"):
                st.write(f"**المادة:** {row['subject']} | **الصف:** {row['grade_level']} | **التاريخ:** {row['visit_date']}")
                st.write(f"**المسؤول العلمي / الموجه:** {row['supervisor_name']} | **النوع:** {row['gender_type']}")
                
                rec_dict = row.to_dict()
                
                sub_tab1, sub_tab2, sub_tab3, sub_tab4 = st.tabs(["📥 تحميل (PDF / Excel)", "✏️ تعديل الاستمارة", "🗑️ حذف الاستمارة", "📂 درايف"])
                
                with sub_tab1:
                    c_btn_pdf, c_btn_xl = st.columns(2)
                    with c_btn_pdf:
                        direct_pdf = generate_direct_pdf_bytes(rec_dict)
                        if direct_pdf:
                            st.download_button(
                                label="📥 تحميل استمارة PDF مباشرة",
                                data=direct_pdf,
                                file_name=f"استمارة_{row['teacher_name']}_{row['visit_date']}.pdf",
                                mime="application/pdf",
                                type="primary",
                                use_container_width=True,
                                key=f"dl_pdf_{eval_id}"
                            )
                        else:
                            st.caption("ℹ️ استخدم زر الطباعة أسفله لحفظ PDF عبر المتصفح.")
                            
                    with c_btn_xl:
                        excel_form_bytes = generate_evaluation_excel_form(rec_dict)
                        st.download_button(
                            label="📊 تحميل استمارة Excel رسمية",
                            data=excel_form_bytes,
                            file_name=f"استمارة_{row['teacher_name']}_{row['visit_date']}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_xl_{eval_id}"
                        )
                        
                    st.markdown("---")
                    st.markdown("#### 🖨️ معاينة الاستمارة الرسمية:")
                    html_preview = get_evaluation_html(rec_dict)
                    components.html(html_preview, height=750, scrolling=True)

                with sub_tab2:
                    e_tname = st.text_input("اسم المدرس", value=row["teacher_name"], key=f"et_{eval_id}")
                    e_subj = st.text_input("المادة", value=row["subject"], key=f"es_{eval_id}")
                    e_topic = st.text_input("موضوع الدرس", value=row["lesson_topic"], key=f"etp_{eval_id}")
                    
                    col_e1, col_e2 = st.columns(2)
                    with col_e1:
                        e_cnt = st.number_input("عدد الطلاب", value=int(row["student_count"]), key=f"ecn_{eval_id}")
                        e_sec = st.text_input("الشعبة", value=row["section"], key=f"esc_{eval_id}")
                    with col_e2:
                        e_status = st.selectbox("الوضع الوظيفي", ["أصيل", "وكيل", "مكلف"], index=["أصيل", "وكيل", "مكلف"].index(row["job_status"]), key=f"est_{eval_id}")
                        e_exp = st.text_input("الخبرة", value=row["experience"], key=f"eex_{eval_id}")
                    
                    st.markdown("**تعديل درجات البنود الـ 25 (قوائم منسدلة):**")
                    curr_scores = json.loads(row["scores_json"]) if isinstance(row["scores_json"], str) else row["scores_json"]
                    updated_scores = {}
                    
                    for it in CRITERIA:
                        val_now = int(curr_scores.get(str(it['id']), it['max']))
                        edit_opts = list(range(it['max'], -1, -1))
                        updated_scores[str(it['id'])] = st.selectbox(
                            f"بند {it['id']}. {it['text']} (الدرجة القصوى: {it['max']}):",
                            options=edit_opts,
                            index=edit_opts.index(val_now) if val_now in edit_opts else 0,
                            key=f"edit_sc_select_{eval_id}_{it['id']}",
                            format_func=lambda x, m=it['max']: f"⭐️ {x} من {m}" if x == m else f"⚪️ {x} من {m}"
                        )
                        
                    e_exc = st.text_area("نقاط التميز", value=row["excellence_points"], key=f"eex2_{eval_id}")
                    e_dev = st.text_area("نقاط التطوير", value=row["dev_points"], key=f"edv_{eval_id}")
                    e_sug = st.text_area("المقترحات", value=row["suggestions"], key=f"esg_{eval_id}")
                    
                    e_files = st.file_uploader("📷 رفع شواهد إضافية", accept_multiple_files=True, key=f"ef_{eval_id}")

                    if st.button("💾 حفظ وتحديث التعديلات", type="primary", use_container_width=True, key=f"btn_edit_{eval_id}"):
                        new_tot = sum(updated_scores.values())
                        new_rat = "ممتاز" if new_tot >= 90 else "جيد جداً" if new_tot >= 80 else "جيد" if new_tot >= 70 else "مقبول" if new_tot >= 50 else "ضعيف"
                        
                        existing_links = str(row["drive_links"]).split(',') if row["drive_links"] else []
                        if e_files:
                            for ef in e_files:
                                link = upload_file_to_drive(ef.getvalue(), ef.name, row["school_name"], date.today(), ef.type)
                                if link: existing_links.append(link)

                        conn = sqlite3.connect("evaluation_system.db")
                        c = conn.cursor()
                        c.execute('''UPDATE evaluations SET 
                            teacher_name=?, subject=?, lesson_topic=?, student_count=?, section=?, job_status=?, experience=?,
                            scores_json=?, total_score=?, rating=?, excellence_points=?, dev_points=?, suggestions=?, drive_links=?
                            WHERE id=?''', (
                                e_tname, e_subj, e_topic, e_cnt, e_sec, e_status, e_exp,
                                json.dumps(updated_scores), new_tot, new_rat, e_exc, e_dev, e_sug, ",".join(existing_links), eval_id
                            ))
                        conn.commit()
                        conn.close()
                        threading.Thread(target=auto_backup_database_to_drive).start()
                        st.success("✅ تم تحديث الاستمارة والنسخة الاحتياطية بنجاح!")
                        st.rerun()

                with sub_tab3:
                    st.error(f"تحذير: سيتم حذف استمارة المدرس ({row['teacher_name']}) في مدرسة ({row['school_name']}) بشكل دائم.")
                    if st.button("🗑️ تأكيد حذف هذه الاستمارة نهائياً", type="primary", use_container_width=True, key=f"confirm_del_btn_{eval_id}"):
                        delete_evaluation_by_id(eval_id)
                        st.warning(f"🗑️ تم حذف استمارة ({row['teacher_name']}) بنجاح!")
                        st.rerun()

                with sub_tab4:
                    if row.get("drive_links"):
                        st.markdown("📂 **الملفات المحفوظة على Google Drive:**")
                        for i, link in enumerate(str(row["drive_links"]).split(','), 1):
                            if link.strip():
                                st.markdown(f"- [عرض الملف {i} على درايف]({link.strip()})")
                    else:
                        st.info("لا توجد ملفات مرفوعة لهذه الاستمارة.")

# -------------------------------------------------------------
# 4. ميزة تتبع ومقارنة تطور أداء المدرس
# -------------------------------------------------------------
elif choice == "📈 تتبع تطور المدرسين":
    st.markdown("### 📈 تتبع ومنحنى تطور أداء المدرس")
    
    conn = sqlite3.connect("evaluation_system.db")
    if st.session_state.user["role"] == "Admin":
        df_all = pd.read_sql_query("SELECT * FROM evaluations", conn)
    else:
        df_all = pd.read_sql_query("SELECT * FROM evaluations WHERE supervisor_name=?", conn, params=(st.session_state.user["name"],))
    conn.close()
    
    if df_all.empty:
        st.info("لا توجد زيارات مسجلة لإجراء التتبع.")
    else:
        unique_teachers = df_all["teacher_name"].dropna().unique().tolist()
        selected_teacher = st.selectbox("اختر اسم المدرس لعرض ملف التطور التتبعي", unique_teachers)
        
        if selected_teacher:
            t_df = df_all[df_all["teacher_name"] == selected_teacher].copy()
            t_df['v_date_dt'] = pd.to_datetime(t_df['visit_date'], errors='coerce')
            t_df = t_df.sort_values(by='v_date_dt', ascending=True)
            
            visits_count = len(t_df)
            
            st.markdown(f"""
            <div class='mobile-card'>
                <div style='font-size: 18px; font-weight: 800; color: #0d5c3a;'>👨‍🏫 المدرس: {selected_teacher}</div>
                <div style='font-size: 14px; color: #64748b;'>المادة: {t_df['subject'].iloc[0]} | عدد الزيارات المسجلة: <b>{visits_count}</b></div>
            </div>
            """, unsafe_allow_html=True)
            
            if visits_count == 1:
                st.info("ℹ️ تم تسجيل زيارة واحدة فقط لهذا المدرس. بمجرد تسجيل زيارة ثانية سيظهر منحنى المقارنة ومؤشرات التطور تلقائياً.")
                single_row = t_df.iloc[0]
                st.write(f"**المدرسة:** {single_row['school_name']} | **التاريخ:** {single_row['visit_date']}")
                st.write(f"**الدرجة الكلية:** {single_row['total_score']} / 100 ({single_row['rating']})")
            else:
                first_visit = t_df.iloc[0]
                latest_visit = t_df.iloc[-1]
                score_diff = latest_visit['total_score'] - first_visit['total_score']
                
                c_m1, c_m2, c_m3 = st.columns(3)
                with c_m1:
                    st.metric("الزيارة الأولى", f"{first_visit['total_score']} / 100", f"{first_visit['visit_date']}")
                with c_m2:
                    st.metric("الزيارة الأحدث", f"{latest_visit['total_score']} / 100", f"{latest_visit['visit_date']}")
                with c_m3:
                    st.metric("فارق التطور", f"{score_diff:+d} درجة", delta_color="normal" if score_diff >= 0 else "inverse")
                
                st.markdown("#### 📊 منحنى التطور عبر الزيارات المتعاقبة:")
                chart_data = t_df[['visit_date', 'total_score']].set_index('visit_date')
                st.line_chart(chart_data)
                
                st.markdown("#### 📋 مقارنة أداء المجالات بين الزيارة الأولى والأحدث:")
                first_scores_dict = json.loads(first_visit['scores_json']) if isinstance(first_visit['scores_json'], str) else first_visit['scores_json']
                latest_scores_dict = json.loads(latest_visit['scores_json']) if isinstance(latest_visit['scores_json'], str) else latest_visit['scores_json']
                
                first_doms = calculate_domain_scores(first_scores_dict)
                latest_doms = calculate_domain_scores(latest_scores_dict)
                
                comp_rows = []
                for d in DOMAINS:
                    s1 = first_doms[d]
                    s2 = latest_doms[d]
                    diff = s2 - s1
                    status_text = "📈 تحسن" if diff > 0 else "📉 تراجع" if diff < 0 else "➖ مستقر"
                    comp_rows.append({
                        "مجال التقييم": d,
                        f"درجة الزيارة 1 ({first_visit['visit_date']})": s1,
                        f"درجة الزيارة الأخيرة ({latest_visit['visit_date']})": s2,
                        "الفارق": f"{diff:+d}",
                        "التقييم": status_text
                    })
                st.dataframe(pd.DataFrame(comp_rows), use_container_width=True)

# -------------------------------------------------------------
# 5. لوحة الإدارة والحماية والنسخ الاحتياطي (Admin)
# -------------------------------------------------------------
elif choice == "⚙️ إدارة النظام والحماية" and st.session_state.user["role"] == "Admin":
    st.markdown("### ⚙️ إدارة النظام والنسخ الاحتياطي والحسابات")
    admin_tab1, admin_tab2, admin_tab3 = st.tabs(["🔒 النسخ الاحتياطي والحماية", "👥 إدارة الموجهين", "🏫 إدارة المدارس"])
    
    with admin_tab1:
        st.markdown("<div class='mobile-card'>", unsafe_allow_html=True)
        st.markdown("#### 🛡️ أمان البيانات والنسخ الاحتياطي السحابي")
        st.write("يقوم النظام تلقائياً برفع نسخة احتياطية من قاعدة البيانات `evaluation_system.db` إلى Google Drive عند كل عملية حفظ أو تعديل أو حذف.")
        
        if st.button("🚀 إنشاء ورفع نسخة احتياطية فورية إلى Google Drive الآن", type="primary", use_container_width=True):
            with st.spinner("جاري أخذ نسخة ورفعها لدرايف..."):
                auto_backup_database_to_drive()
                st.success("✅ تم رفع نسخة احتياطية حديثة إلى مجلد (النسخ_الاحتياطية_للنظام) في Google Drive بنجاح!")
                
        if os.path.exists("evaluation_system.db"):
            with open("evaluation_system.db", "rb") as f_db:
                st.download_button(
                    label="💾 تنزيل ملف قاعدة البيانات المحلي (.db)",
                    data=f_db.read(),
                    file_name=f"evaluation_system_backup_{date.today()}.db",
                    mime="application/x-sqlite3",
                    use_container_width=True
                )
        st.markdown("</div>", unsafe_allow_html=True)

    with admin_tab2:
        conn = sqlite3.connect("evaluation_system.db")
        users_df = pd.read_sql_query("SELECT id, username, full_name, specialization, role FROM users", conn)
        conn.close()
        st.dataframe(users_df, use_container_width=True)
        
        st.markdown("#### ✏️ تعديل بيانات موجه")
        selected_u_id = st.selectbox("اختر المستخدم للتعديل", users_df["id"].tolist(), format_func=lambda x: f"ID {x}: {users_df[users_df['id'] == x]['full_name'].values[0]}")
        if selected_u_id:
            conn = sqlite3.connect("evaluation_system.db")
            c = conn.cursor()
            c.execute("SELECT id, username, password, full_name, specialization, role FROM users WHERE id=?", (selected_u_id,))
            u_row = c.fetchone()
            conn.close()
            with st.form(f"edit_u_form_{selected_u_id}"):
                e_fname = st.text_input("الاسم الكامل", value=u_row[3])
                e_uname = st.text_input("اسم المستخدم", value=u_row[1])
                e_pword = st.text_input("كلمة المرور", value=u_row[2])
                e_spec = st.text_input("التخصص", value=u_row[4])
                e_role = st.selectbox("الدور", ["Supervisor", "Admin"], index=0 if u_row[5] == "Supervisor" else 1)
                
                sb1, sb2 = st.columns(2)
                with sb1: save_u = st.form_submit_button("💾 حفظ التعديل", use_container_width=True)
                with sb2: del_u = st.form_submit_button("🗑️ حذف الحساب", use_container_width=True)
                if save_u:
                    conn = sqlite3.connect("evaluation_system.db")
                    c = conn.cursor()
                    c.execute("UPDATE users SET username=?, password=?, full_name=?, specialization=?, role=? WHERE id=?", (e_uname, e_pword, e_fname, e_spec, e_role, selected_u_id))
                    conn.commit()
                    conn.close()
                    threading.Thread(target=auto_backup_database_to_drive).start()
                    st.success("تم التحديث بنجاح")
                    st.rerun()

    with admin_tab3:
        conn = sqlite3.connect("evaluation_system.db")
        schools_all = pd.read_sql_query("SELECT id, name, gender, location FROM schools", conn)
        conn.close()
        st.dataframe(schools_all, use_container_width=True)
        
        with st.expander("➕ إضافة مدرسة جديدة"):
            with st.form("add_sc_form"):
                ns_name = st.text_input("اسم المدرسة")
                ns_gen = st.selectbox("النوع", ["ذكور", "إناث"])
                ns_loc = st.text_input("الموقع الجغرافي")
                if st.form_submit_button("➕ إضافة المدرسة", use_container_width=True):
                    if ns_name:
                        conn = sqlite3.connect("evaluation_system.db")
                        c = conn.cursor()
                        c.execute("INSERT INTO schools (name, gender, location) VALUES (?, ?, ?)", (ns_name, ns_gen, ns_loc))
                        conn.commit()
                        conn.close()
                        threading.Thread(target=auto_backup_database_to_drive).start()
                        st.success("تمت إضافة المدرسة بنجاح")
                        st.rerun()
